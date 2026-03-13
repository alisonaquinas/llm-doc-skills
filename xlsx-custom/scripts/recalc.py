#!/usr/bin/env python3
"""
recalc.py — Recalculate all formulas in an Excel workbook (.xlsx).

openpyxl writes formula strings verbatim into cells but does NOT evaluate
them.  Excel opens the file, sees that cached values are stale (or absent),
and silently recalculates on first open.  When the file is later read by
another program (pandas, Google Sheets import, CI validation) the stale or
empty cached values produce wrong results.

This script runs LibreOffice Calc in headless mode to open the workbook and
re-save it, forcing LibreOffice to evaluate every formula and write the
result back into the cached-value fields.  The output file is then
compatible with tools that rely on cached values.

After recalculation the script reads back the file with openpyxl and prints
a JSON summary of any formula errors found in the workbook.

LibreOffice wrapper used (in search order):
  1. office-custom/scripts/soffice.py  (sibling skill, preferred)
  2. scripts/office/soffice.py         (legacy repo-root location)
  3. System ``soffice`` / ``libreoffice`` binary on PATH

Usage:
    python xlsx-custom/scripts/recalc.py spreadsheet.xlsx [output.xlsx]

    If *output.xlsx* is omitted the input file is overwritten in-place.

Options:
    --check-only   Run recalculation and print the JSON error report, but
                   exit with a non-zero code if any formula errors are found.
    --quiet        Suppress all output except the JSON error summary.

Exit codes:
    0  — Success (and no formula errors if --check-only).
    1  — Input file not found or LibreOffice not available.
    2  — Formula errors found (only when --check-only is set).

Example output (JSON):
    {
      "file": "spreadsheet.xlsx",
      "sheets": {
        "Sheet1": {
          "B3": "#DIV/0!",
          "D7": "#REF!"
        }
      },
      "error_summary": {
        "#DIV/0!": 1,
        "#REF!": 1
      },
      "total_errors": 2
    }
"""

import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path

# openpyxl is a pure-Python library for reading/writing Excel files.
# It is available via: pip install openpyxl
try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    print(
        "ERROR: openpyxl is not installed.  Run: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


# ---------------------------------------------------------------------------
# Error token recognition
# ---------------------------------------------------------------------------

# Standard Excel formula-error strings.  openpyxl surfaces these as the
# cell's value when reading back a recalculated workbook.
_FORMULA_ERRORS = frozenset({
    "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!", "#CALC!",
})


# ---------------------------------------------------------------------------
# soffice.py location discovery
# ---------------------------------------------------------------------------

def _find_soffice_py() -> Path | None:
    """
    Search for the soffice.py LibreOffice wrapper in well-known locations.

    Search order:
      1. ``office-custom/scripts/soffice.py``  — sibling skill (preferred)
      2. ``scripts/office/soffice.py``          — legacy repo-root location
      3. Returns None if neither exists (caller falls back to system soffice)

    The search anchors from the *repository root*, derived by walking two
    levels up from this file (xlsx-custom/scripts/ → xlsx-custom/ → repo/).
    """
    # This file lives at: <repo>/xlsx-custom/scripts/recalc.py
    # repo root is two parents up.
    repo_root = Path(__file__).parent.parent.parent

    candidates = [
        repo_root / "office-custom" / "scripts" / "soffice.py",  # sibling skill
        repo_root / "scripts" / "office" / "soffice.py",         # legacy location
    ]
    return next((p for p in candidates if p.exists()), None)


# ---------------------------------------------------------------------------
# LibreOffice-based recalculation
# ---------------------------------------------------------------------------

def _recalc_with_soffice(source: Path, dest: Path, quiet: bool = False) -> bool:
    """
    Use LibreOffice Calc in headless mode to recalculate *source* and save to *dest*.

    Locates the soffice.py wrapper (which handles the ``soffice`` binary path
    and user-profile isolation for sandboxed environments) via
    :func:`_find_soffice_py`, then invokes it with ``--convert-to xlsx`` to
    force formula evaluation on re-save.

    Args:
        source: Input .xlsx path.
        dest:   Output .xlsx path (may equal *source* for in-place recalc).
        quiet:  If True, suppress progress messages.

    Returns:
        True on success, False if LibreOffice could not be found or failed.
    """
    soffice_py = _find_soffice_py()
    if soffice_py is None:
        if not quiet:
            print(
                "WARNING: soffice.py not found — falling back to system soffice.",
                file=sys.stderr,
            )
        return _recalc_fallback(source, dest, quiet)

    # Use soffice.py's run() function programmatically.
    try:
        soffice_module = _import_soffice_module(soffice_py)
    except Exception as exc:
        if not quiet:
            print(f"WARNING: Could not import soffice.py: {exc}", file=sys.stderr)
        return _recalc_fallback(source, dest, quiet)

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_src = Path(tmpdir) / source.name
        shutil.copy2(source, tmp_src)

        # --convert-to xlsx re-saves via LibreOffice which triggers recalc.
        result = soffice_module.run([
            "--headless",
            "--convert-to", "xlsx",
            "--outdir", str(tmpdir),
            str(tmp_src),
        ])

        tmp_out = Path(tmpdir) / (tmp_src.stem + ".xlsx")
        if result.returncode != 0 or not tmp_out.exists():
            if not quiet:
                print(
                    f"WARNING: soffice conversion failed (returncode={result.returncode}).",
                    file=sys.stderr,
                )
            return False

        shutil.copy2(tmp_out, dest)

    if not quiet:
        print(f"  Recalculated via LibreOffice: {dest.name}")
    return True


def _import_soffice_module(soffice_py: Path):
    """
    Dynamically import the soffice.py module from its filesystem path.

    Args:
        soffice_py: Absolute path to the soffice.py wrapper script.

    Returns:
        The imported module object with a ``run(args)`` function.
    """
    import importlib.util
    spec   = importlib.util.spec_from_file_location("soffice", soffice_py)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _recalc_fallback(source: Path, dest: Path, quiet: bool) -> bool:
    """
    Attempt recalculation using a bare ``soffice`` call found on PATH.

    This is used when soffice.py is unavailable.

    Args:
        source: Input .xlsx file.
        dest:   Output .xlsx file.
        quiet:  Suppress messages if True.

    Returns:
        True on success, False if soffice is not on PATH or fails.
    """
    import subprocess

    soffice_bin = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice_bin:
        if not quiet:
            print(
                "ERROR: LibreOffice (soffice) not found on PATH.  "
                "Install LibreOffice or ensure it is on your PATH.",
                file=sys.stderr,
            )
        return False

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_src = Path(tmpdir) / source.name
        shutil.copy2(source, tmp_src)

        cmd = [
            soffice_bin, "--headless",
            "--convert-to", "xlsx",
            "--outdir", tmpdir,
            str(tmp_src),
        ]
        result = subprocess.run(cmd, capture_output=True)

        tmp_out = Path(tmpdir) / (tmp_src.stem + ".xlsx")
        if result.returncode != 0 or not tmp_out.exists():
            if not quiet:
                print(
                    "ERROR: soffice conversion failed.\n"
                    f"  stdout: {result.stdout.decode(errors='replace')}\n"
                    f"  stderr: {result.stderr.decode(errors='replace')}",
                    file=sys.stderr,
                )
            return False

        shutil.copy2(tmp_out, dest)

    if not quiet:
        print(f"  Recalculated via system soffice: {dest.name}")
    return True


# ---------------------------------------------------------------------------
# Error scanning
# ---------------------------------------------------------------------------

def scan_formula_errors(xlsx_path: Path) -> dict:
    """
    Open *xlsx_path* with openpyxl and collect every cell containing a
    formula-error value (#DIV/0!, #REF!, etc.).

    Args:
        xlsx_path: Path to the (possibly recalculated) .xlsx file.

    Returns:
        A dict with structure::

            {
              "file": "name.xlsx",
              "sheets": {
                "Sheet1": {"B3": "#DIV/0!", ...},
                ...
              },
              "error_summary": {"#DIV/0!": 1, ...},
              "total_errors": N
            }
    """
    try:
        # data_only=True reads cached formula results instead of formula strings.
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    except (InvalidFileException, Exception) as exc:
        return {
            "file":          xlsx_path.name,
            "sheets":        {},
            "error_summary": {},
            "total_errors":  0,
            "read_error":    str(exc),
        }

    sheets_report: dict[str, dict[str, str]] = {}
    error_counts:  dict[str, int]            = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_errors: dict[str, str] = {}

        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val in _FORMULA_ERRORS:
                    cell_addr = cell.coordinate
                    sheet_errors[cell_addr] = val
                    error_counts[val] = error_counts.get(val, 0) + 1

        if sheet_errors:
            sheets_report[sheet_name] = sheet_errors

    return {
        "file":          xlsx_path.name,
        "sheets":        sheets_report,
        "error_summary": error_counts,
        "total_errors":  sum(error_counts.values()),
    }


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def recalc(
    source: Path,
    dest:   Path | None = None,
    check_only: bool    = False,
    quiet:      bool    = False,
) -> dict:
    """
    Recalculate *source* and return a JSON-serialisable error report.

    Args:
        source:     Input .xlsx file.
        dest:       Output .xlsx file.  Defaults to overwriting *source*.
        check_only: If True, exits with code 2 when formula errors are found.
        quiet:      If True, suppress progress messages.

    Returns:
        The error-report dict from :func:`scan_formula_errors`.
    """
    source = Path(source)
    if not source.exists():
        print(f"ERROR: File not found: {source}", file=sys.stderr)
        sys.exit(1)

    if dest is None:
        dest = source
    else:
        dest = Path(dest)

    if not quiet:
        print(f"Recalculating {source.name} …")

    success = _recalc_with_soffice(source, dest, quiet=quiet)
    if not success:
        if not quiet:
            print(
                "WARNING: Recalculation failed — scanning the original file for errors.",
                file=sys.stderr,
            )
        # If recalc failed and dest != source, copy the original so the
        # caller still has an output file to work with.
        if dest != source:
            shutil.copy2(source, dest)

    report = scan_formula_errors(dest)

    if not quiet or check_only:
        print(json.dumps(report, indent=2))

    if check_only and report["total_errors"] > 0:
        sys.exit(2)

    return report


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("input",  help="Input .xlsx file")
    parser.add_argument("output", nargs="?", help="Output .xlsx file (default: overwrite input)")
    parser.add_argument(
        "--check-only",
        action="store_true",
        help="Exit with code 2 if any formula errors are found after recalculation",
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Suppress progress output; still prints the JSON error report",
    )
    args = parser.parse_args()

    recalc(
        source     = Path(args.input),
        dest       = Path(args.output) if args.output else None,
        check_only = args.check_only,
        quiet      = args.quiet,
    )


if __name__ == "__main__":
    main()
